from datetime import datetime
from io import BytesIO
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook

# Add current folder to sys.path
sys.path.insert(0, str(Path(__file__).parent))

from engine import (
    CLEAN_HEADERS,
    DATE_NUMBER_FORMAT,
    CURRENCY_NUMBER_FORMAT,
    TEXT_NUMBER_FORMAT,
    extract_pichincha_transactions,
    generate_clean_excel,
    process_pichincha_file,
)


def _create_sample_pichincha_raw_workbook() -> bytes:
    """Create a mock Pichincha export with 2-row layout, offset document/beneficiary, and Spanish amounts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    headers = {
        4: "Fecha",
        5: "Concepto",
        9: "Nro. Documento",
        11: "Tipo",
        14: "Beneficiario",
        15: "Monto",
        16: "Saldo",
    }
    for col, name in headers.items():
        ws.cell(6, col).value = name

    # Transaction 1: Credit / Income
    ws.cell(7, 4).value = "2026-6-30, 1:26 AM"
    ws.cell(7, 5).value = "INTERES A SU FAVOR"
    ws.cell(7, 11).value = "Crédito"
    ws.cell(7, 15).value = "$0,13"
    ws.cell(7, 16).value = "$14.342,22"
    ws.cell(8, 8).value = "881757353"  # Offset document number on row below

    # Transaction 2: Debit / Expense
    ws.cell(11, 4).value = "2026-6-30, 9:22 PM"
    ws.cell(11, 5).value = "COMPRA SUPERMERCADO"
    ws.cell(11, 11).value = "Débito"
    ws.cell(11, 15).value = "-$20,00"
    ws.cell(11, 16).value = "$14.342,09"
    ws.cell(12, 8).value = "78631587"
    ws.cell(12, 14).value = "******4374"  # Offset beneficiary on row below

    # Transaction 3: Standard ISO date format
    ws.cell(15, 4).value = "2026-07-01 10:30"
    ws.cell(15, 5).value = "TRANSFERENCIA DIRECTA"
    ws.cell(15, 11).value = "Crédito"
    ws.cell(15, 15).value = "$150,50"
    ws.cell(15, 16).value = "$14.492,59"
    ws.cell(16, 9).value = "000998877"
    ws.cell(16, 14).value = "JUAN PEREZ"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_extract_pichincha_transactions():
    raw_bytes = _create_sample_pichincha_raw_workbook()
    records, reports = extract_pichincha_transactions(raw_bytes)

    assert len(records) == 3
    assert list(records[0].keys()) == CLEAN_HEADERS

    # Verify Row 1 (Credit)
    assert records[0]["Fecha"] == datetime(2026, 6, 30, 1, 26)
    assert records[0]["Concepto"] == "INTERES A SU FAVOR"
    assert records[0]["Nro. Documento"] == "881757353"
    assert records[0]["Tipo"] == "Crédito"
    assert records[0]["Monto"] == 0.13
    assert records[0]["Saldo"] == 14342.22

    # Verify Row 2 (Debit)
    assert records[1]["Fecha"] == datetime(2026, 6, 30, 21, 22)
    assert records[1]["Concepto"] == "COMPRA SUPERMERCADO"
    assert records[1]["Nro. Documento"] == "78631587"
    assert records[1]["Tipo"] == "Débito"
    assert records[1]["Beneficiario"] == "******4374"
    assert records[1]["Monto"] == -20.0
    assert records[1]["Saldo"] == 14342.09

    # Verify Row 3 (Credit with beneficiary)
    assert records[2]["Fecha"] == datetime(2026, 7, 1, 10, 30)
    assert records[2]["Concepto"] == "TRANSFERENCIA DIRECTA"
    assert records[2]["Nro. Documento"] == "000998877"
    assert records[2]["Tipo"] == "Crédito"
    assert records[2]["Beneficiario"] == "JUAN PEREZ"
    assert records[2]["Monto"] == 150.50
    assert records[2]["Saldo"] == 14492.59

    # Check report statistics
    assert len(reports) == 1
    assert reports[0]["transactions"] == 3
    assert reports[0]["income_count"] == 2
    assert reports[0]["expense_count"] == 1
    assert reports[0]["total_income"] == 150.63
    assert reports[0]["total_expenses"] == 20.0


def test_generate_clean_excel_table_and_formats():
    raw_bytes = _create_sample_pichincha_raw_workbook()
    records, _ = extract_pichincha_transactions(raw_bytes)
    clean_bytes = generate_clean_excel(records, sheet_name="Movimientos")

    wb = load_workbook(BytesIO(clean_bytes), data_only=False)
    ws = wb["Movimientos"]

    # Verify dimensions: 1 header row + 3 data rows = 4 rows, 7 columns
    assert ws.max_row == 4
    assert ws.max_column == 7

    # Verify Headers
    header_values = [ws.cell(row=1, column=c).value for c in range(1, 8)]
    assert header_values == CLEAN_HEADERS

    # Verify Excel Table object is added
    assert len(ws.tables) == 1
    table = list(ws.tables.values())[0]
    assert table.displayName == "MovimientosPichincha"
    assert table.ref == "A1:G4"
    assert table.tableStyleInfo.name == "TableStyleMedium9"
    assert table.tableStyleInfo.showRowStripes is True

    # Verify Frozen pane at A2
    assert ws.freeze_panes == "A2"

    # Verify Date number format is 'dd-mm-yyyy hh:mm'
    assert ws["A2"].number_format == DATE_NUMBER_FORMAT
    assert ws["A2"].value == datetime(2026, 6, 30, 1, 26)

    # Verify Text format for Document number
    assert ws["C2"].number_format == TEXT_NUMBER_FORMAT
    assert ws["C2"].value == "881757353"

    # Verify Currency format for Monto and Saldo
    assert ws["F2"].number_format == CURRENCY_NUMBER_FORMAT
    assert ws["F2"].value == 0.13
    assert ws["G2"].number_format == CURRENCY_NUMBER_FORMAT
    assert ws["G2"].value == 14342.22

    # Verify Debit amount is negative
    assert ws["F3"].value == -20.0


def test_process_pichincha_file_roundtrip(tmp_path):
    raw_bytes = _create_sample_pichincha_raw_workbook()
    src_file = tmp_path / "estado_cuenta_pichincha.xlsx"
    src_file.write_bytes(raw_bytes)

    out_file, stats = process_pichincha_file(src_file)

    assert out_file.exists()
    assert out_file.name == "estado_cuenta_pichincha_filtrable.xlsx"
    assert stats["transactions"] == 3
    assert stats["total_income"] == 150.63

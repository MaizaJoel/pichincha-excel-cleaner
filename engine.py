"""Banco Pichincha Excel Cleaner Engine - 100% Local & Offline.

This module parses messy Banco Pichincha Excel exports (which place metadata,
document numbers, and beneficiaries across alternating/offset rows) and produces
a normalized, official Excel Table (ListObject) with filter headers, custom date
formatting (dd-mm-yyyy hh:mm), currency styling, and auto-adjusted columns.

Zero external processes or cloud dependencies required.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
import re
import unicodedata
from typing import BinaryIO

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


CLEAN_HEADERS = [
    "Fecha",
    "Concepto",
    "Nro. Documento",
    "Tipo",
    "Beneficiario",
    "Monto",
    "Saldo",
]

# Custom formatting constants
DATE_NUMBER_FORMAT = "dd-mm-yyyy hh:mm"
CURRENCY_NUMBER_FORMAT = '$#,##0.00;[Red]-$#,##0.00;"-"'
TEXT_NUMBER_FORMAT = "@"


def _clean_text(value: object) -> str:
    """Normalize whitespace and remove hidden Excel carriage return codes."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("_x000D_", " ")).strip()


def _normalize_header(value: object) -> str:
    """ASCII normalize text for case/accent-insensitive header lookup."""
    normalized = unicodedata.normalize("NFKD", _clean_text(value)).encode("ascii", "ignore").decode().upper()
    return re.sub(r"[^A-Z0-9]+", "", normalized)


def _parse_money(value: object, default: Decimal | None = Decimal("0")) -> Decimal | None:
    """Parse Ecuadorian currency strings (e.g. '$14.342,22' or '-$20,00') reliably."""
    if value is None:
        return default
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return default
    text = _clean_text(value).replace("$", "").replace(" ", "")
    if not text:
        return default
    negative = text.startswith("-") or (text.startswith("(") and text.endswith(")")) or text.endswith("-")
    text = text.strip("()-+")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
    elif re.search(r",\d{1,2}$", text):
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        parsed = Decimal(text)
        return -parsed if negative else parsed
    except InvalidOperation:
        return default


def _parse_datetime(value: object) -> datetime | None:
    """Parse Ecuadorian Spanish datetimes from Pichincha exports."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = _clean_text(value).replace(",", " ")
    text = re.sub(r"\s+", " ", text).strip()
    
    # ISO-like match: 2026-06-30 01:26 AM / 2026-6-30, 1:26 AM
    iso_match = re.match(r"^(?P<year>\d{4})-(?P<month>\d{1,2})-(?P<day>\d{1,2})(?:\s+(?P<hour>\d{1,2}):(?P<minute>\d{2})(?:\s*(?P<ampm>AM|PM))?)?", text, re.I)
    if iso_match:
        try:
            g = iso_match.groupdict()
            hour = int(g["hour"] or 0)
            if g["ampm"]:
                hour = hour % 12 + (12 if g["ampm"].upper() == "PM" else 0)
            return datetime(int(g["year"]), int(g["month"]), int(g["day"]), hour, int(g["minute"] or 0))
        except (ValueError, TypeError):
            pass

    # Spanish month names map
    spanish_months = {
        "ene": "jan", "feb": "feb", "mar": "mar", "abr": "apr", "may": "may", "jun": "jun",
        "jul": "jul", "ago": "aug", "sep": "sep", "set": "sep", "oct": "oct", "nov": "nov", "dic": "dec",
    }
    for spanish, english in spanish_months.items():
        text = re.sub(rf"(?i)(?<=-)({spanish})(?=\b)", english, text)

    formats = (
        "%d/%m/%Y %H:%M", "%d/%m/%Y %I:%M %p", "%d/%m/%Y",
        "%d-%m-%Y %H:%M", "%d-%m-%Y %I:%M %p", "%d-%m-%Y",
        "%Y-%m-%d %H:%M", "%Y-%m-%d %I:%M %p", "%Y-%m-%d",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def extract_pichincha_transactions(file_or_bytes: str | Path | bytes | BinaryIO) -> tuple[list[dict], list[dict]]:
    """Extract and merge 2-row Banco Pichincha transaction statements into clean dict records.
    
    Returns:
        records: List of dictionaries matching CLEAN_HEADERS
        summary_stats: List of sheet-level summaries (total rows, removed rows, income count, etc.)
    """
    if isinstance(file_or_bytes, bytes):
        wb = load_workbook(BytesIO(file_or_bytes), read_only=True, data_only=True)
    elif isinstance(file_or_bytes, (str, Path)):
        wb = load_workbook(str(file_or_bytes), read_only=True, data_only=True)
    else:
        wb = load_workbook(file_or_bytes, read_only=True, data_only=True)

    records: list[dict] = []
    summary_report: list[dict] = []

    for worksheet in wb.worksheets:
        raw_rows = list(worksheet.iter_rows(values_only=True))
        if not raw_rows:
            continue

        # Find header row
        header_idx = next(
            (
                i for i, row in enumerate(raw_rows)
                if any(_normalize_header(v) == "FECHA" for v in row)
                and any(_normalize_header(v) in {"CONCEPTO", "DESCRIPCION"} for v in row)
            ),
            None,
        )
        if header_idx is None:
            continue

        header_row = raw_rows[header_idx]

        def get_col_index(*labels: str) -> int | None:
            wanted = {_normalize_header(l) for l in labels}
            return next((i for i, val in enumerate(header_row) if _normalize_header(val) in wanted), None)

        date_col = get_col_index("Fecha")
        desc_col = get_col_index("Concepto", "Descripcion")
        doc_col = get_col_index("Nro. Documento", "Documento", "Nro Documento")
        type_col = get_col_index("Tipo")
        ben_col = get_col_index("Beneficiario")
        amount_col = get_col_index("Monto", "Valor")
        balance_col = get_col_index("Saldo")

        def look_ahead_metadata(row_num: int, col_idx: int | None, kind: str) -> str:
            if col_idx is None:
                return ""
            candidate_cols = [col_idx]
            if kind == "document":
                if col_idx > 0:
                    candidate_cols.append(col_idx - 1)
                candidate_cols.append(col_idx + 1)

            for offset in range(0, 3):
                target_r = row_num + offset
                if target_r >= len(raw_rows):
                    break
                candidate_row = raw_rows[target_r]
                for c in candidate_cols:
                    if c >= len(candidate_row):
                        continue
                    val = _clean_text(candidate_row[c])
                    if not val or val.lower().startswith("total"):
                        continue
                    if kind == "document":
                        if re.fullmatch(r"\d{5,}", val):
                            return val
                    elif val not in {"Monto", "Saldo", "Tipo"}:
                        return val
            return ""

        parsed_sheet_records: list[dict] = []
        income_count = 0
        expense_count = 0
        total_income = Decimal("0")
        total_expenses = Decimal("0")

        for r_idx in range(header_idx + 1, len(raw_rows)):
            row = raw_rows[r_idx]
            if date_col is None or not row[date_col] or desc_col is None or not row[desc_col]:
                continue

            typ = _clean_text(row[type_col]).lower() if type_col is not None else ""
            raw_amount = row[amount_col] if amount_col is not None else None
            amount = _parse_money(raw_amount)
            if not amount:
                continue

            dt = _parse_datetime(row[date_col])
            raw_desc = _clean_text(row[desc_col])
            document = look_ahead_metadata(r_idx, doc_col, "document")
            beneficiary = look_ahead_metadata(r_idx, ben_col, "beneficiary")
            raw_balance = row[balance_col] if balance_col is not None else None
            balance = _parse_money(raw_balance, default=None)

            if "cred" in typ:
                movement_type = "Crédito"
                final_amount = abs(amount)
                income_count += 1
                total_income += final_amount
            elif "deb" in typ:
                movement_type = "Débito"
                final_amount = -abs(amount)
                expense_count += 1
                total_expenses += abs(amount)
            else:
                if amount > 0:
                    movement_type = "Crédito"
                    final_amount = amount
                    income_count += 1
                    total_income += final_amount
                else:
                    movement_type = "Débito"
                    final_amount = amount
                    expense_count += 1
                    total_expenses += abs(final_amount)

            item = {
                "Fecha": dt,
                "Concepto": raw_desc,
                "Nro. Documento": document,
                "Tipo": movement_type,
                "Beneficiario": beneficiary,
                "Monto": float(final_amount),
                "Saldo": float(balance) if balance is not None else None,
            }
            parsed_sheet_records.append(item)
            records.append(item)

        output_rows = len(parsed_sheet_records) + 1
        summary_report.append({
            "sheet": worksheet.title,
            "rows_before": len(raw_rows),
            "rows_removed": max(len(raw_rows) - output_rows, 0),
            "rows_after": output_rows,
            "transactions": len(parsed_sheet_records),
            "income_count": income_count,
            "expense_count": expense_count,
            "total_income": float(total_income),
            "total_expenses": float(total_expenses),
            "net_flow": float(total_income - total_expenses),
        })

    if not records:
        raise ValueError(
            "No se encontraron movimientos con el formato de Banco Pichincha. "
            "Verifica que el archivo contenga las columnas: Fecha, Concepto, Tipo, Monto, Saldo."
        )

    return records, summary_report


def generate_clean_excel(records: list[dict], sheet_name: str = "Movimientos") -> bytes:
    """Generate a high-quality Excel file with a formatted Excel Table (ListObject).
    
    Includes:
    - Active Header Auto-Filter buttons
    - Table styling (TableStyleMedium9 - elegant header with alternating banded rows)
    - Requested date format 'dd-mm-yyyy hh:mm'
    - Document number '@' format (text)
    - Currency format '$#,##0.00;[Red]-$#,##0.00;\"-\"'
    - Frozen top header row
    - Dynamic auto-calculated column widths
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    ws.append(CLEAN_HEADERS)

    # Style Header Row
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx in range(1, len(CLEAN_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for record in records:
        row_values = [
            record.get("Fecha"),
            record.get("Concepto"),
            record.get("Nro. Documento"),
            record.get("Tipo"),
            record.get("Beneficiario"),
            record.get("Monto"),
            record.get("Saldo"),
        ]
        ws.append(row_values)

    total_rows = len(records) + 1

    # Apply cell formatting across data rows
    date_alignment = Alignment(horizontal="center", vertical="center")
    doc_alignment = Alignment(horizontal="center", vertical="center")
    type_alignment = Alignment(horizontal="center", vertical="center")
    text_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")

    for row_idx in range(2, total_rows + 1):
        # Column 1: Fecha (Date)
        cell_a = ws.cell(row=row_idx, column=1)
        cell_a.number_format = DATE_NUMBER_FORMAT
        cell_a.alignment = date_alignment

        # Column 2: Concepto (Description)
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.alignment = text_alignment

        # Column 3: Nro. Documento (Document Number)
        cell_c = ws.cell(row=row_idx, column=3)
        cell_c.number_format = TEXT_NUMBER_FORMAT
        cell_c.alignment = doc_alignment

        # Column 4: Tipo (Movement Type)
        cell_d = ws.cell(row=row_idx, column=4)
        cell_d.alignment = type_alignment

        # Column 5: Beneficiario (Beneficiary)
        cell_e = ws.cell(row=row_idx, column=5)
        cell_e.alignment = text_alignment

        # Column 6: Monto (Amount)
        cell_f = ws.cell(row=row_idx, column=6)
        cell_f.number_format = CURRENCY_NUMBER_FORMAT
        cell_f.alignment = number_alignment

        # Column 7: Saldo (Balance)
        cell_g = ws.cell(row=row_idx, column=7)
        cell_g.number_format = CURRENCY_NUMBER_FORMAT
        cell_g.alignment = number_alignment

    # Create official Excel Table (ListObject)
    if records:
        table_ref = f"A1:G{total_rows}"
        table = Table(displayName="MovimientosPichincha", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Freeze Header Pane
    ws.freeze_panes = "A2"

    # Set optimal Column Widths
    column_widths = {
        1: 20,  # Fecha
        2: 45,  # Concepto
        3: 16,  # Nro. Documento
        4: 12,  # Tipo
        5: 22,  # Beneficiario
        6: 16,  # Monto
        7: 16,  # Saldo
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Export to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def process_pichincha_file(input_path: str | Path, output_path: str | Path | None = None) -> tuple[Path, dict]:
    """Convenience helper to read an input file and write the cleaned table directly to disk."""
    src = Path(input_path)
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {src}")

    records, report = extract_pichincha_transactions(src)
    xlsx_bytes = generate_clean_excel(records)

    if output_path is None:
        dst = src.parent / f"{src.stem}_filtrable.xlsx"
    else:
        dst = Path(output_path)

    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_bytes(xlsx_bytes)

    stats = report[0] if report else {}
    return dst, stats
